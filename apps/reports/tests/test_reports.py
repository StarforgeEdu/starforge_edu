"""Lane B (reports) test suite — DAY-4 D4-LB "Tests required".

Covers: the six generators against factory data (incl. ai_usage consuming Lane
A's selector, mocked), role-visibility matrix, teacher cohort scoping enforced in
the selector, the signed-URL build flow (mocked boto3 helpers), schedule
exactly-once within the hour, cross-tenant isolation + query-count, and the PDF/Excel render path (skipped when
weasyprint/openpyxl are absent — the data/collect layer is asserted instead).
"""

from __future__ import annotations

from datetime import date, datetime, timedelta
from typing import Any

import pytest
from django.test import override_settings
from django.utils import timezone
from django_tenants.utils import schema_context

from apps.attendance.models import AttendanceRecord
from apps.cohorts.tests.factories import CohortFactory
from apps.org.tests.factories import BranchFactory
from apps.reports import selectors, services
from apps.reports.generators import get_generator
from apps.reports.models import Report, ReportRun, ReportSchedule
from apps.reports.serializers import ReportRunReadSerializer
from apps.schedule.models import Lesson
from apps.schedule.tests.factories import TermFactory
from apps.students.models import StudentProfile
from apps.students.tests.factories import StudentProfileFactory
from apps.teachers.tests.factories import TeacherProfileFactory
from core.exceptions import TenantContextMissing, ThrottledException, ValidationException
from core.permissions import Role

pytestmark = pytest.mark.django_db


# --------------------------------------------------------------------------- #
# Optional-render-lib probes (mirror the academics transcript skip)
# --------------------------------------------------------------------------- #
try:  # pragma: no cover - import probe
    import weasyprint  # noqa: F401

    _HAS_WEASYPRINT = True
except Exception:
    _HAS_WEASYPRINT = False

try:  # pragma: no cover - import probe
    import openpyxl  # noqa: F401

    _HAS_OPENPYXL = True
except Exception:
    _HAS_OPENPYXL = False


# --------------------------------------------------------------------------- #
# Builders
# --------------------------------------------------------------------------- #
def _make_lesson(*, branch, teacher, cohort, starts_at=None):
    term = TermFactory()
    starts_at = starts_at or (timezone.now() - timedelta(hours=1))
    return Lesson.objects.create(
        term=term,
        cohort=cohort,
        teacher=teacher,
        title="Algebra",
        starts_at=starts_at,
        ends_at=starts_at + timedelta(hours=1),
    )


def _custom_account_type(*, slug: str, grants: set[str], account_kind: str = "staff"):
    from apps.access.models import AccountType, AccountTypePermission

    account_type = AccountType.objects.create(
        name=slug.replace("-", " ").title(),
        slug=slug,
        account_kind=account_kind,
    )
    AccountTypePermission.objects.bulk_create(
        [
            AccountTypePermission(account_type=account_type, permission=permission)
            for permission in sorted(grants)
        ]
    )
    return account_type


def _assign_account_type(*, user, branch, account_type, department=None):
    from apps.users.models import RoleMembership

    membership = RoleMembership.objects.create(
        user=user,
        branch=branch,
        department=department,
        account_type=account_type,
        role=account_type.compatibility_role,
    )
    user.refresh_from_db()
    return membership


# --------------------------------------------------------------------------- #
# Library seed (D4-LB-1)
# --------------------------------------------------------------------------- #
def test_six_library_rows_seeded(tenant_a):
    with schema_context(tenant_a.schema_name):
        keys = set(Report.objects.values_list("key", flat=True))
    assert keys == {"enrollment", "attendance", "grades", "finance", "ai_usage", "storage_usage"}


# --------------------------------------------------------------------------- #
# Generator correctness x6 (D4-LB-3)
# --------------------------------------------------------------------------- #
def test_enrollment_generator_counts(tenant_a, user_in):
    with schema_context(tenant_a.schema_name):
        director = user_in(tenant_a, roles=[Role.DIRECTOR])
        branch = BranchFactory()
        cohort = CohortFactory(branch=branch)
        for _ in range(3):
            s: Any = StudentProfileFactory(branch=branch, status=StudentProfile.Status.ACTIVE)
            s.current_cohort = cohort
            s.save(update_fields=["current_cohort"])
        # A lead student must NOT count (only enrolled/active are seats).
        StudentProfileFactory(branch=branch, status=StudentProfile.Status.LEAD)

        data = get_generator("enrollment").collect({}, user=director, roles={Role.DIRECTOR})
    assert data["total"] == 3
    assert data["by_status"]["active"] == 3
    assert len(data["rows"]) == 3


def test_attendance_generator_status_counts(tenant_a, user_in):
    with schema_context(tenant_a.schema_name):
        director = user_in(tenant_a, roles=[Role.DIRECTOR])
        branch = BranchFactory()
        cohort = CohortFactory(branch=branch)
        teacher = TeacherProfileFactory(branch=branch)
        lesson = _make_lesson(branch=branch, teacher=teacher, cohort=cohort)
        students = [StudentProfileFactory(branch=branch) for _ in range(2)]
        AttendanceRecord.objects.create(student=students[0], lesson=lesson, status="present")
        AttendanceRecord.objects.create(student=students[1], lesson=lesson, status="absent")

        data = get_generator("attendance").collect({}, user=director, roles={Role.DIRECTOR})
    assert data["total"] == 2
    assert data["by_status"] == {"present": 1, "absent": 1}


def test_report_row_cap_refuses_oversized_report(tenant_a, user_in, monkeypatch):
    """R2-06: a report generator must refuse (not OOM) an oversized result set — it
    materializes every row into a list + an in-memory doc. With the cap lowered, an
    unfiltered full-scope collect raises report_too_large instead of loading all rows."""
    from apps.reports.generators import base
    from core.exceptions import ValidationException

    monkeypatch.setattr(base, "MAX_REPORT_ROWS", 1)
    with schema_context(tenant_a.schema_name):
        director = user_in(tenant_a, roles=[Role.DIRECTOR])
        branch = BranchFactory()
        cohort = CohortFactory(branch=branch)
        teacher = TeacherProfileFactory(branch=branch)
        lesson = _make_lesson(branch=branch, teacher=teacher, cohort=cohort)
        for _ in range(2):  # 2 rows > cap of 1
            AttendanceRecord.objects.create(
                student=StudentProfileFactory(branch=branch), lesson=lesson, status="present"
            )
        with pytest.raises(ValidationException) as exc:
            get_generator("attendance").collect({}, user=director, roles={Role.DIRECTOR})
        assert exc.value.code == "report_too_large"


def test_grades_generator_published_only(tenant_a, user_in):
    from apps.academics.tests.factories import GradeFactory

    with schema_context(tenant_a.schema_name):
        director = user_in(tenant_a, roles=[Role.DIRECTOR])
        GradeFactory(is_published=True)
        GradeFactory(is_published=False)  # excluded unless include_unpublished

        published = get_generator("grades").collect({}, user=director, roles={Role.DIRECTOR})
        both = get_generator("grades").collect(
            {"include_unpublished": True}, user=director, roles={Role.DIRECTOR}
        )
    assert published["total"] == 1
    assert both["total"] == 2


def test_finance_generator_totals(tenant_a, user_in):
    from decimal import Decimal

    from apps.finance.models import Invoice
    from core.historical_scope import ScopeAttributionStatus

    with schema_context(tenant_a.schema_name):
        accountant = user_in(tenant_a, roles=[Role.ACCOUNTANT])
        branch_id = accountant.role_memberships.values_list("branch_id", flat=True).get()
        student: Any = StudentProfileFactory(branch_id=branch_id)
        Invoice.objects.create(
            number="INV-2026-000001",
            student=student,
            branch_at_issue_id=branch_id,
            attribution_status=ScopeAttributionStatus.CAPTURED,
            status=Invoice.Status.ISSUED,
            total_uzs=Decimal("100000.00"),
            issue_date=date(2026, 6, 1),
        )
        Invoice.objects.create(
            number="INV-2026-000002",
            student=student,
            branch_at_issue_id=branch_id,
            attribution_status=ScopeAttributionStatus.CAPTURED,
            status=Invoice.Status.PAID,
            total_uzs=Decimal("50000.00"),
            issue_date=date(2026, 6, 1),
        )
        data = get_generator("finance").collect({}, user=accountant, roles={Role.ACCOUNTANT})
    assert data["total_invoices"] == 2
    assert data["total_billed_uzs"] == "150000.00"
    # Only the open (issued) invoice with no allocations is outstanding.
    assert data["outstanding_uzs"] == "100000.00"


def test_ai_usage_generator_consumes_lane_a_selector(tenant_a, user_in, monkeypatch):
    # ai_usage calls apps.ai.selectors.tokens_consumed (Lane A interface). Stub it.
    import apps.reports.generators.ai_usage as ai_mod

    monkeypatch.setattr(ai_mod, "_tokens_consumed", lambda start, end: 4242)
    with schema_context(tenant_a.schema_name):
        director = user_in(tenant_a, roles=[Role.DIRECTOR])
        data = get_generator("ai_usage").collect({"month": "2026-06"}, user=director, roles={Role.DIRECTOR})
    assert data["tokens_consumed"] == 4242
    assert data["month"] == "2026-06"


def test_ai_usage_returns_authoritative_integer(tenant_a, user_in):
    with schema_context(tenant_a.schema_name):
        director = user_in(tenant_a, roles=[Role.DIRECTOR])
        data = get_generator("ai_usage").collect({}, user=director, roles={Role.DIRECTOR})
    assert isinstance(data["tokens_consumed"], int)


def test_ai_usage_selector_failure_is_not_reported_as_zero(tenant_a, user_in, monkeypatch):
    import apps.reports.generators.ai_usage as ai_mod

    def unavailable(*args, **kwargs):
        raise RuntimeError("usage source unavailable")

    monkeypatch.setattr(ai_mod, "_tokens_consumed", unavailable)
    with schema_context(tenant_a.schema_name):
        director = user_in(tenant_a, roles=[Role.DIRECTOR])
        with pytest.raises(RuntimeError, match="usage source unavailable"):
            get_generator("ai_usage").collect({}, user=director, roles={Role.DIRECTOR})


def test_storage_usage_generator_sums_clean_bytes(tenant_a, user_in):
    from apps.content.models import ContentLibrary, Folder, LessonFile

    with schema_context(tenant_a.schema_name):
        director = user_in(tenant_a, roles=[Role.DIRECTOR])
        lib = ContentLibrary.objects.create(name="Lib")
        folder = Folder.objects.create(library=lib, name="F")
        LessonFile.objects.create(
            folder=folder,
            title="a",
            s3_key="k/a",
            content_type="application/pdf",
            size_bytes=1000,
            status=LessonFile.Status.CLEAN,
        )
        LessonFile.objects.create(
            folder=folder,
            title="b",
            s3_key="k/b",
            content_type="application/pdf",
            size_bytes=500,
            status=LessonFile.Status.CLEAN,
        )
        # A pending file is NOT counted.
        LessonFile.objects.create(
            folder=folder,
            title="c",
            s3_key="k/c",
            content_type="application/pdf",
            size_bytes=9999,
            status=LessonFile.Status.PENDING,
        )
        data = get_generator("storage_usage").collect({}, user=director, roles={Role.DIRECTOR})
    assert data["total_bytes"] == 1500
    assert data["total_files"] == 2


# --------------------------------------------------------------------------- #
# Teacher cohort scoping enforced in the selector (D4-LB-5)
# --------------------------------------------------------------------------- #
def test_teacher_attendance_scoped_to_own_cohorts(tenant_a, user_in):
    with schema_context(tenant_a.schema_name):
        branch = BranchFactory()
        teacher_user = user_in(tenant_a, roles=[Role.TEACHER], branch=branch)
        teacher = TeacherProfileFactory(user=teacher_user, branch=branch)
        # Own cohort (teacher is primary).
        own = CohortFactory(branch=branch, name="Own", primary_teacher=teacher)
        own_lesson = _make_lesson(branch=branch, teacher=teacher, cohort=own)
        own_student = StudentProfileFactory(branch=branch)
        AttendanceRecord.objects.create(student=own_student, lesson=own_lesson, status="present")
        # Foreign cohort (another teacher).
        other_teacher = TeacherProfileFactory(branch=branch)
        foreign = CohortFactory(branch=branch, name="Foreign", primary_teacher=other_teacher)
        foreign_lesson = _make_lesson(branch=branch, teacher=other_teacher, cohort=foreign)
        foreign_student = StudentProfileFactory(branch=branch)
        AttendanceRecord.objects.create(student=foreign_student, lesson=foreign_lesson, status="absent")

        data = get_generator("attendance").collect({}, user=teacher_user, roles={Role.TEACHER})
    # Only the teacher's own cohort row is present; the foreign cohort is absent.
    assert data["total"] == 1
    assert data["rows"][0]["cohort"] == "Own"


def test_teacher_enrollment_scoped_to_own_cohorts(tenant_a, user_in):
    with schema_context(tenant_a.schema_name):
        branch = BranchFactory()
        teacher_user = user_in(tenant_a, roles=[Role.TEACHER], branch=branch)
        teacher = TeacherProfileFactory(user=teacher_user, branch=branch)
        own = CohortFactory(branch=branch, name="Own", primary_teacher=teacher)
        s1: Any = StudentProfileFactory(branch=branch, status=StudentProfile.Status.ACTIVE)
        s1.current_cohort = own
        s1.save(update_fields=["current_cohort"])
        # Foreign-cohort student (no membership/ownership by this teacher).
        foreign = CohortFactory(branch=branch, name="Foreign")
        s2: Any = StudentProfileFactory(branch=branch, status=StudentProfile.Status.ACTIVE)
        s2.current_cohort = foreign
        s2.save(update_fields=["current_cohort"])

        data = get_generator("enrollment").collect({}, user=teacher_user, roles={Role.TEACHER})
    assert data["total"] == 1
    assert data["rows"][0]["cohort"] == "Own"


# --------------------------------------------------------------------------- #
# Role-visibility matrix (D4-LB-5)
# --------------------------------------------------------------------------- #
def test_library_visibility_by_role(tenant_a, user_in):
    with schema_context(tenant_a.schema_name):
        director = user_in(tenant_a, roles=[Role.DIRECTOR])
        accountant = user_in(tenant_a, roles=[Role.ACCOUNTANT])
        teacher = user_in(tenant_a, roles=[Role.TEACHER])

        dir_keys = set(
            selectors.scoped_reports(user=director, roles={Role.DIRECTOR}).values_list("key", flat=True)
        )
        acc_keys = set(
            selectors.scoped_reports(user=accountant, roles={Role.ACCOUNTANT}).values_list("key", flat=True)
        )
        tea_keys = set(
            selectors.scoped_reports(user=teacher, roles={Role.TEACHER}).values_list("key", flat=True)
        )
    assert dir_keys == {"enrollment", "attendance", "grades", "finance", "ai_usage", "storage_usage"}
    assert acc_keys == {"finance"}
    assert tea_keys == {"enrollment", "attendance", "grades"}


def test_stale_director_role_cannot_globalize_local_compound_report_grant(tenant_a, user_in):
    """A malformed compatibility role is neither a grant nor a scope boundary."""
    from apps.reports.authorization import snapshot_from_params
    from apps.users.models import RoleMembership

    with schema_context(tenant_a.schema_name):
        local = BranchFactory(name="Report Local", slug="report-local")
        remote = BranchFactory(name="Report Remote", slug="report-remote")
        local_type = _custom_account_type(
            slug="local-report-students",
            grants={"reports:read", "reports:write", "students:read"},
        )
        empty_type = _custom_account_type(slug="empty-report-role", grants=set())
        actor = user_in(tenant_a)
        _assign_account_type(user=actor, branch=local, account_type=local_type)
        stale = _assign_account_type(user=actor, branch=remote, account_type=empty_type)
        RoleMembership.objects.filter(pk=stale.pk).update(role=Role.DIRECTOR)
        actor.refresh_from_db()

        roles = services._live_roles(actor)
        keys = set(selectors.scoped_reports(user=actor, roles=roles).values_list("key", flat=True))
        run = services.create_report_run(
            report_key="enrollment",
            fmt="pdf",
            params={},
            requested_by=actor,
            roles=roles,
        )
        scope = snapshot_from_params(run.params)

    assert keys == {"enrollment"}
    assert scope is not None
    assert scope.organization is False
    assert scope.branch_ids == (local.pk,)


def test_split_report_and_domain_grants_cannot_be_composed_across_departments(tenant_a, user_in):
    from apps.org.tests.factories import DepartmentFactory
    from core.exceptions import PermissionException

    with schema_context(tenant_a.schema_name):
        branch = BranchFactory(name="Split Reports", slug="split-reports")
        reports_department = DepartmentFactory(
            branch=branch, name="Reports Department", slug="reports-department"
        )
        domain_department = DepartmentFactory(
            branch=branch, name="Domain Department", slug="domain-department"
        )
        reports_type = _custom_account_type(
            slug="reports-only",
            grants={"reports:read", "reports:write"},
        )
        domain_type = _custom_account_type(
            slug="students-only",
            grants={"students:read"},
        )
        actor = user_in(tenant_a)
        _assign_account_type(
            user=actor,
            branch=branch,
            department=reports_department,
            account_type=reports_type,
        )
        _assign_account_type(
            user=actor,
            branch=branch,
            department=domain_department,
            account_type=domain_type,
        )
        roles = services._live_roles(actor)

        assert not selectors.scoped_reports(user=actor, roles=roles).filter(key="enrollment").exists()
        with pytest.raises(PermissionException) as exc:
            services.create_report_run(
                report_key="enrollment",
                fmt="pdf",
                params={},
                requested_by=actor,
                roles=roles,
            )
    assert exc.value.code == "report_forbidden"


def test_split_grants_at_same_exact_department_remain_additive(tenant_a, user_in):
    from apps.org.tests.factories import DepartmentFactory
    from apps.reports.authorization import snapshot_from_params

    with schema_context(tenant_a.schema_name):
        branch = BranchFactory(name="Additive Reports", slug="additive-reports")
        department = DepartmentFactory(branch=branch, name="Additive Department", slug="additive-department")
        reports_type = _custom_account_type(
            slug="additive-reports-only",
            grants={"reports:read", "reports:write"},
        )
        domain_type = _custom_account_type(
            slug="additive-students-only",
            grants={"students:read"},
        )
        actor = user_in(tenant_a)
        for account_type in (reports_type, domain_type):
            _assign_account_type(
                user=actor,
                branch=branch,
                department=department,
                account_type=account_type,
            )
        roles = services._live_roles(actor)
        run = services.create_report_run(
            report_key="enrollment",
            fmt="pdf",
            params={},
            requested_by=actor,
            roles=roles,
        )
        scope = snapshot_from_params(run.params)

    assert scope is not None
    assert scope.department_keys == (f"{branch.pk}:{department.pk}",)


def test_canonical_teacher_without_cohort_param_never_receives_branch_wide_rows(tenant_a, user_in):
    with schema_context(tenant_a.schema_name):
        branch = BranchFactory(name="Teacher Report Branch", slug="teacher-report-branch")
        teacher_user = user_in(tenant_a, roles=[Role.TEACHER], branch=branch)
        teacher = TeacherProfileFactory(user=teacher_user, branch=branch)
        own = CohortFactory(branch=branch, name="Canonical Own", primary_teacher=teacher)
        own_lesson = _make_lesson(branch=branch, teacher=teacher, cohort=own)
        AttendanceRecord.objects.create(
            student=StudentProfileFactory(branch=branch),
            lesson=own_lesson,
            status="present",
        )
        other_teacher = TeacherProfileFactory(branch=branch)
        foreign = CohortFactory(branch=branch, name="Canonical Foreign", primary_teacher=other_teacher)
        foreign_lesson = _make_lesson(branch=branch, teacher=other_teacher, cohort=foreign)
        AttendanceRecord.objects.create(
            student=StudentProfileFactory(branch=branch),
            lesson=foreign_lesson,
            status="absent",
        )

        roles = services._live_roles(teacher_user)
        data = get_generator("attendance").collect({}, user=teacher_user, roles=roles)
        run = services.create_report_run(
            report_key="attendance",
            fmt="pdf",
            params={},
            requested_by=teacher_user,
            roles=roles,
        )

    assert data["total"] == 1
    assert data["rows"][0]["cohort"] == "Canonical Own"
    assert run.params["_scope_branch_ids"] == []
    assert run.params["_scope_teacher_keys"]
    assert run.params["_scope_teacher_cohort_ids"] == [own.pk]


def test_revocation_after_queue_blocks_worker_before_render(tenant_a, user_in, monkeypatch):
    from apps.reports.generators.base import ReportGenerator
    from core.exceptions import PermissionException

    rendered = False

    def render(*args, **kwargs):
        nonlocal rendered
        rendered = True
        return b"must-not-render"

    monkeypatch.setattr(ReportGenerator, "render", render)
    with schema_context(tenant_a.schema_name):
        branch = BranchFactory(name="Revoked Reports", slug="revoked-reports")
        report_type = _custom_account_type(
            slug="revocable-reports",
            grants={"reports:read", "reports:write", "students:read"},
        )
        actor = user_in(tenant_a)
        membership = _assign_account_type(user=actor, branch=branch, account_type=report_type)
        run = services.create_report_run(
            report_key="enrollment",
            fmt="pdf",
            params={},
            requested_by=actor,
            roles=services._live_roles(actor),
        )
        membership.revoked_at = timezone.now()
        membership.save(update_fields=["revoked_at"])

        with pytest.raises(PermissionException) as exc:
            services.build_report_run(run.pk)

    assert exc.value.code == "report_forbidden"
    assert rendered is False


def test_teacher_assignment_removal_hides_run_and_blocks_queued_worker(tenant_a, user_in):
    from core.exceptions import PermissionException

    with schema_context(tenant_a.schema_name):
        branch = BranchFactory(name="Teacher Revocation", slug="teacher-revocation")
        teacher_user = user_in(tenant_a, roles=[Role.TEACHER], branch=branch)
        teacher = TeacherProfileFactory(user=teacher_user, branch=branch)
        cohort = CohortFactory(branch=branch, name="Former Teacher Cohort", primary_teacher=teacher)
        roles = services._live_roles(teacher_user)
        run = services.create_report_run(
            report_key="enrollment",
            fmt="pdf",
            params={},
            requested_by=teacher_user,
            roles=roles,
        )
        assert selectors.scoped_runs(user=teacher_user, roles=roles).filter(pk=run.pk).exists()

        cohort.primary_teacher = None
        cohort.save(update_fields=["primary_teacher"])
        current_roles = services._live_roles(teacher_user)
        assert not selectors.scoped_runs(user=teacher_user, roles=current_roles).filter(pk=run.pk).exists()
        with pytest.raises(PermissionException) as exc:
            services.build_report_run(run.pk)

    assert exc.value.code == "report_forbidden"


def test_multi_scope_run_and_schedule_disappear_when_one_scope_is_revoked(tenant_a, user_in):
    with schema_context(tenant_a.schema_name):
        first = BranchFactory(name="Report Union A", slug="report-union-a")
        second = BranchFactory(name="Report Union B", slug="report-union-b")
        report_type = _custom_account_type(
            slug="union-report-type",
            grants={"reports:read", "reports:write", "students:read"},
        )
        actor = user_in(tenant_a)
        _assign_account_type(user=actor, branch=first, account_type=report_type)
        second_membership = _assign_account_type(user=actor, branch=second, account_type=report_type)
        roles = services._live_roles(actor)
        run = services.create_report_run(
            report_key="enrollment",
            fmt="pdf",
            params={},
            requested_by=actor,
            roles=roles,
        )
        schedule = services.create_schedule(
            report_key="enrollment",
            created_by=actor,
            roles=roles,
            cadence=ReportSchedule.Cadence.WEEKLY,
            weekday=0,
        )
        assert selectors.scoped_runs(user=actor, roles=roles).filter(pk=run.pk).exists()
        assert selectors.scoped_schedules(user=actor, roles=roles).filter(pk=schedule.pk).exists()

        second_membership.revoked_at = timezone.now()
        second_membership.save(update_fields=["revoked_at"])
        current_roles = services._live_roles(actor)

        assert not selectors.scoped_runs(user=actor, roles=current_roles).filter(pk=run.pk).exists()
        assert not selectors.scoped_schedules(user=actor, roles=current_roles).filter(pk=schedule.pk).exists()


def test_cross_scope_run_and_schedule_detail_ids_return_not_found(tenant_a, user_in, as_user):
    with schema_context(tenant_a.schema_name):
        local = BranchFactory(name="Run Scope Local", slug="run-scope-local")
        remote = BranchFactory(name="Run Scope Remote", slug="run-scope-remote")
        report_type = _custom_account_type(
            slug="detail-report-type",
            grants={"reports:read", "reports:write", "students:read"},
        )
        creator = user_in(tenant_a)
        _assign_account_type(user=creator, branch=local, account_type=report_type)
        outsider = user_in(tenant_a)
        _assign_account_type(user=outsider, branch=remote, account_type=report_type)
        roles = services._live_roles(creator)
        run = services.create_report_run(
            report_key="enrollment",
            fmt="pdf",
            params={"branch_id": local.pk},
            requested_by=creator,
            roles=roles,
        )
        schedule = services.create_schedule(
            report_key="enrollment",
            created_by=creator,
            roles=roles,
            cadence=ReportSchedule.Cadence.WEEKLY,
            weekday=0,
            params={"branch_id": local.pk},
        )

    client = as_user(tenant_a, outsider)
    assert client.get(f"/api/v1/reports/runs/{run.pk}/").status_code == 404
    assert client.get(f"/api/v1/reports/schedules/{schedule.pk}/").status_code == 404


def test_schedule_update_cannot_borrow_write_grant_from_another_branch(tenant_a, user_in, as_user):
    with schema_context(tenant_a.schema_name):
        local = BranchFactory(name="Schedule Read Local", slug="schedule-read-local")
        remote = BranchFactory(name="Schedule Write Remote", slug="schedule-write-remote")
        local_reader = _custom_account_type(
            slug="local-schedule-reader",
            grants={"reports:read", "students:read"},
        )
        remote_writer = _custom_account_type(
            slug="remote-schedule-writer",
            grants={"reports:write", "students:read"},
        )
        actor = user_in(tenant_a)
        _assign_account_type(user=actor, branch=local, account_type=local_reader)
        _assign_account_type(user=actor, branch=remote, account_type=remote_writer)

        director = user_in(tenant_a, roles=[Role.DIRECTOR], branch=local)
        schedule = services.create_schedule(
            report_key="enrollment",
            created_by=director,
            roles=services._live_roles(director),
            cadence=ReportSchedule.Cadence.WEEKLY,
            weekday=0,
            params={"branch_id": local.pk},
        )

    client = as_user(tenant_a, actor)
    assert client.get(f"/api/v1/reports/schedules/{schedule.pk}/").status_code == 200
    response = client.patch(
        f"/api/v1/reports/schedules/{schedule.pk}/",
        {"params": {"branch_id": remote.pk}},
        format="json",
    )
    assert response.status_code == 403


def test_accountant_cannot_run_grades(tenant_a, user_in):
    from core.exceptions import PermissionException

    with schema_context(tenant_a.schema_name):
        accountant = user_in(tenant_a, roles=[Role.ACCOUNTANT])
        with pytest.raises(PermissionException) as exc:
            services.create_report_run(
                report_key="grades",
                fmt="pdf",
                params={},
                requested_by=accountant,
                roles={Role.ACCOUNTANT},
            )
    assert exc.value.code == "report_forbidden"


def test_accountant_can_run_finance(tenant_a, user_in):
    with schema_context(tenant_a.schema_name):
        accountant = user_in(tenant_a, roles=[Role.ACCOUNTANT])
        run = services.create_report_run(
            report_key="finance",
            fmt="xlsx",
            params={},
            requested_by=accountant,
            roles={Role.ACCOUNTANT},
        )
    assert run.status == ReportRun.Status.QUEUED
    assert run.format == "xlsx"


def test_runs_are_object_scoped_by_owner_and_branch(tenant_a, user_in):
    with schema_context(tenant_a.schema_name):
        own_branch = BranchFactory()
        foreign_branch = BranchFactory()
        director = user_in(tenant_a, roles=[Role.DIRECTOR], branch=own_branch)
        own_hod = user_in(tenant_a, roles=[Role.HEAD_OF_DEPT], branch=own_branch)
        foreign_hod = user_in(tenant_a, roles=[Role.HEAD_OF_DEPT], branch=foreign_branch)
        run = services.create_report_run(
            report_key="enrollment",
            fmt="pdf",
            params={"branch_id": own_branch.pk},
            requested_by=director,
            roles={Role.DIRECTOR},
        )
        assert selectors.scoped_runs(user=own_hod, roles={Role.HEAD_OF_DEPT}).filter(pk=run.pk).exists()
        assert (
            not selectors.scoped_runs(user=foreign_hod, roles={Role.HEAD_OF_DEPT}).filter(pk=run.pk).exists()
        )


@override_settings(REPORT_MAX_ACTIVE_PER_USER=1)
def test_report_admission_dedupes_before_enforcing_queue_cap(tenant_a, user_in):
    with schema_context(tenant_a.schema_name):
        accountant = user_in(tenant_a, roles=[Role.ACCOUNTANT])
        first = services.create_report_run(
            report_key="finance",
            fmt="xlsx",
            params={},
            requested_by=accountant,
            roles={Role.ACCOUNTANT},
        )
        duplicate = services.create_report_run(
            report_key="finance",
            fmt="xlsx",
            params={},
            requested_by=accountant,
            roles={Role.ACCOUNTANT},
        )
        assert duplicate.pk == first.pk
        with pytest.raises(ThrottledException) as exc:
            services.create_report_run(
                report_key="finance",
                fmt="pdf",
                params={},
                requested_by=accountant,
                roles={Role.ACCOUNTANT},
            )
        assert exc.value.code == "report_user_queue_full"


def test_schedule_recipients_must_be_active_and_in_branch_scope(tenant_a, user_in):
    with schema_context(tenant_a.schema_name):
        branch = BranchFactory()
        foreign = BranchFactory()
        hod = user_in(tenant_a, roles=[Role.HEAD_OF_DEPT], branch=branch)
        outsider = user_in(tenant_a, roles=[Role.TEACHER], branch=foreign)
        with pytest.raises(ValidationException) as exc:
            services.create_schedule(
                report_key="enrollment",
                created_by=hod,
                roles={Role.HEAD_OF_DEPT},
                cadence=ReportSchedule.Cadence.WEEKLY,
                weekday=0,
                recipient_ids=[outsider.pk],
            )
    assert exc.value.code == "invalid_recipients"


def test_schedule_recipient_must_cover_complete_exact_department_scope(tenant_a, user_in):
    from apps.org.tests.factories import DepartmentFactory

    with schema_context(tenant_a.schema_name):
        branch = BranchFactory(name="Recipient Branch", slug="recipient-branch")
        allowed_department = DepartmentFactory(
            branch=branch, name="Allowed Recipient Department", slug="allowed-recipient"
        )
        other_department = DepartmentFactory(
            branch=branch, name="Other Recipient Department", slug="other-recipient"
        )
        writer_type = _custom_account_type(
            slug="department-report-writer",
            grants={"reports:read", "reports:write", "students:read"},
        )
        reader_type = _custom_account_type(
            slug="department-report-reader",
            grants={"reports:read", "students:read"},
        )
        creator = user_in(tenant_a)
        _assign_account_type(
            user=creator,
            branch=branch,
            department=allowed_department,
            account_type=writer_type,
        )
        allowed = user_in(tenant_a)
        _assign_account_type(
            user=allowed,
            branch=branch,
            department=allowed_department,
            account_type=reader_type,
        )
        outside = user_in(tenant_a)
        _assign_account_type(
            user=outside,
            branch=branch,
            department=other_department,
            account_type=reader_type,
        )
        roles = services._live_roles(creator)

        schedule = services.create_schedule(
            report_key="enrollment",
            created_by=creator,
            roles=roles,
            cadence=ReportSchedule.Cadence.WEEKLY,
            weekday=0,
            recipient_ids=[allowed.pk],
        )
        assert schedule.recipient_ids == [allowed.pk]

        with pytest.raises(ValidationException) as exc:
            services.create_schedule(
                report_key="enrollment",
                created_by=creator,
                roles=roles,
                cadence=ReportSchedule.Cadence.WEEKLY,
                weekday=0,
                recipient_ids=[outside.pk],
            )
    assert exc.value.code == "invalid_recipients"


def test_tenant_wide_schedule_cannot_send_download_to_branch_only_recipient(tenant_a, user_in):
    with schema_context(tenant_a.schema_name):
        branch = BranchFactory(name="Organization Report", slug="organization-report")
        director = user_in(tenant_a, roles=[Role.DIRECTOR], branch=branch)
        branch_manager = user_in(tenant_a, roles=[Role.HEAD_OF_DEPT], branch=branch)

        with pytest.raises(ValidationException) as exc:
            services.create_schedule(
                report_key="enrollment",
                created_by=director,
                roles=services._live_roles(director),
                cadence=ReportSchedule.Cadence.WEEKLY,
                weekday=0,
                recipient_ids=[branch_manager.pk],
            )
    assert exc.value.code == "invalid_recipients"


# --------------------------------------------------------------------------- #
# Build flow: render → S3 → presign → notifications.dispatch (D4-LB-4)
# --------------------------------------------------------------------------- #
def _patch_s3(monkeypatch):
    """Patch the s3_client helpers used by the report service onto a recorder."""
    from infrastructure.storage import s3_client

    store: dict[str, bytes] = {}

    def upload_bytes(key, data, *, content_type="application/octet-stream"):
        store[key] = data
        return key

    def presign_download(key, *, expires_in=600):
        return f"memory://get/{key}"

    monkeypatch.setattr(s3_client, "upload_bytes", upload_bytes)
    monkeypatch.setattr(s3_client, "presign_download", presign_download)
    return store


def test_build_report_run_flow(tenant_a, user_in, monkeypatch):
    store = _patch_s3(monkeypatch)
    # Render to deterministic bytes regardless of weasyprint/openpyxl availability.
    from apps.reports.generators.base import ReportGenerator

    monkeypatch.setattr(ReportGenerator, "render", lambda self, data, fmt, *, locale="uz": b"%PDF-FAKE")
    dispatched: list[dict] = []
    import apps.reports.services as svc

    def fake_dispatch(**kwargs):
        dispatched.append(kwargs)
        return None

    monkeypatch.setattr("apps.notifications.services.dispatch", fake_dispatch)

    with schema_context(tenant_a.schema_name):
        director = user_in(tenant_a, roles=[Role.DIRECTOR])
        run = ReportRun.objects.create(
            report=Report.objects.get(key="enrollment"),
            requested_by=director,
            format="pdf",
            status=ReportRun.Status.QUEUED,
        )
        key = svc.build_report_run(run.pk)
        run.refresh_from_db()

    assert run.status == ReportRun.Status.DONE
    assert key == f"{tenant_a.schema_name}/reports/{run.pk}.pdf"
    assert run.s3_key == key
    assert run.file_bytes == len(b"%PDF-FAKE")
    assert store[key] == b"%PDF-FAKE"
    # Delivery went through notifications.dispatch (never email directly).
    assert len(dispatched) == 1
    assert dispatched[0]["event_type"] == "report.ready"
    assert dispatched[0]["recipient_id"] == director.pk
    assert dispatched[0]["context"]["download_url"].startswith("memory://get/")


def test_build_report_records_real_notification(tenant_a, user_in, monkeypatch):
    """End-to-end through the REAL notifications.dispatch: a Notification row is
    recorded carrying the report.ready event (the EOD demo's step-3 assertion)."""
    _patch_s3(monkeypatch)
    from apps.notifications.models import Notification
    from apps.reports.generators.base import ReportGenerator

    monkeypatch.setattr(ReportGenerator, "render", lambda self, data, fmt, *, locale="uz": b"%PDF-FAKE")
    with schema_context(tenant_a.schema_name):
        director = user_in(tenant_a, roles=[Role.DIRECTOR])
        run = ReportRun.objects.create(
            report=Report.objects.get(key="enrollment"),
            requested_by=director,
            format="pdf",
            status=ReportRun.Status.QUEUED,
        )
        services.build_report_run(run.pk)
        notif = Notification.objects.filter(user=director, event_type="report.ready").first()
    assert notif is not None
    assert notif.data["run_id"] == run.pk


def test_build_report_idempotent_skips_done(tenant_a, user_in, monkeypatch):
    _patch_s3(monkeypatch)
    import apps.reports.services as svc

    monkeypatch.setattr("apps.notifications.services.dispatch", lambda **kw: None)
    with schema_context(tenant_a.schema_name):
        director = user_in(tenant_a, roles=[Role.DIRECTOR])
        run = ReportRun.objects.create(
            report=Report.objects.get(key="enrollment"),
            requested_by=director,
            format="pdf",
            status=ReportRun.Status.DONE,
            s3_key="already/done.pdf",
        )
        result = svc.build_report_run(run.pk)
    # A terminal row never re-renders, but a non-server-derived legacy key is
    # not trusted or returned as a downloadable object reference.
    assert result is None


def test_presign_run_rejects_forged_cross_tenant_or_cross_report_key(tenant_a, user_in, monkeypatch):
    from infrastructure.storage import s3_client

    seen: list[str] = []
    monkeypatch.setattr(
        s3_client,
        "presign_download",
        lambda key, *, expires_in=600: seen.append(key) or f"memory://{key}",
    )
    with schema_context(tenant_a.schema_name):
        director = user_in(tenant_a, roles=[Role.DIRECTOR])
        run = ReportRun.objects.create(
            report=Report.objects.get(key="enrollment"),
            requested_by=director,
            format="pdf",
            status=ReportRun.Status.DONE,
        )
        expected = f"{tenant_a.schema_name}/reports/{run.pk}.pdf"
        run.s3_key = expected
        assert services.presign_run(run) == f"memory://{expected}"

        for forged in (
            f"other_tenant/reports/{run.pk}.pdf",
            f"{tenant_a.schema_name}/reports/{run.pk + 1}.pdf",
            f"{tenant_a.schema_name}/transcripts/{run.pk}.pdf",
            f"{tenant_a.schema_name}/reports/{run.pk}.xlsx",
        ):
            run.s3_key = forged
            assert services.presign_run(run) is None

    assert seen == [expected]


def test_build_report_marks_failed(tenant_a, user_in, monkeypatch):
    import apps.reports.services as svc
    from apps.reports.generators.base import ReportGenerator

    def boom(self, data, fmt, *, locale="uz"):
        raise RuntimeError("render exploded")

    monkeypatch.setattr(ReportGenerator, "render", boom)
    _patch_s3(monkeypatch)
    with schema_context(tenant_a.schema_name):
        director = user_in(tenant_a, roles=[Role.DIRECTOR])
        run = ReportRun.objects.create(
            report=Report.objects.get(key="enrollment"),
            requested_by=director,
            format="pdf",
            status=ReportRun.Status.QUEUED,
        )
        with pytest.raises(RuntimeError):
            svc.build_report_run(run.pk)
        svc.mark_run_failed(run.pk, RuntimeError("render exploded"))
        run.refresh_from_db()
        assert run.status == ReportRun.Status.FAILED
        assert run.error == "report_generation_failed"
        serialized = ReportRunReadSerializer(run).data
        assert serialized["error"] == "report_generation_failed"
        assert "render exploded" not in str(serialized)

        run.error = "s3://operator:secret@private/report.pdf"
        assert ReportRunReadSerializer(run).data["error"] == "report_generation_failed"

    # A detached tenant model must not lazy-load ``report.key`` from the public
    # schema.  The serializer fails closed before it can issue that query.
    with pytest.raises(TenantContextMissing):
        _ = ReportRunReadSerializer(run).data


# --------------------------------------------------------------------------- #
# Schedule exactly-once within the cadence window (D4-LB-6)
# --------------------------------------------------------------------------- #
def test_schedule_due_fires_exactly_once(tenant_a, user_in):
    with schema_context(tenant_a.schema_name):
        director = user_in(tenant_a, roles=[Role.DIRECTOR])
        report = Report.objects.get(key="enrollment")
        # Choose a `now` whose weekday/hour match the schedule anchor.
        now = timezone.localtime(timezone.now())
        sched = ReportSchedule.objects.create(
            report=report,
            created_by=director,
            cadence=ReportSchedule.Cadence.WEEKLY,
            weekday=now.weekday(),
            hour=now.hour,
            format="pdf",
        )
        fired = services.run_due_schedules(now=now)
        again = services.run_due_schedules(now=now)
        sched.refresh_from_db()
        run_count = ReportRun.objects.filter(report=report).count()
    assert fired == 1
    assert again == 0  # last_run_at guard suppresses the second scan
    assert run_count == 1
    assert sched.last_run_at is not None


def test_schedule_not_due_wrong_hour(tenant_a, user_in):
    with schema_context(tenant_a.schema_name):
        director = user_in(tenant_a, roles=[Role.DIRECTOR])
        now = timezone.localtime(timezone.now())
        ReportSchedule.objects.create(
            report=Report.objects.get(key="enrollment"),
            created_by=director,
            cadence=ReportSchedule.Cadence.WEEKLY,
            weekday=now.weekday(),
            hour=(now.hour + 1) % 24,  # not this hour
            format="pdf",
        )
        fired = services.run_due_schedules(now=now)
    assert fired == 0


def test_monthly_schedule_fires_on_short_month_last_day(tenant_a, user_in):
    """A monthly schedule for day_of_month=31 must still fire in February (clamped
    to the last day) instead of being silently skipped."""
    with schema_context(tenant_a.schema_name):
        director = user_in(tenant_a, roles=[Role.DIRECTOR])
        sched = ReportSchedule.objects.create(
            report=Report.objects.get(key="enrollment"),
            created_by=director,
            cadence=ReportSchedule.Cadence.MONTHLY,
            day_of_month=31,
            hour=9,
            format="pdf",
        )
        # 2026 is not a leap year → Feb has 28 days; day 31 clamps to the 28th.
        feb_last = timezone.make_aware(datetime(2026, 2, 28, 9, 0))
        feb_other = timezone.make_aware(datetime(2026, 2, 27, 9, 0))
        assert services.schedule_is_due(sched, now=feb_last) is True
        assert services.schedule_is_due(sched, now=feb_other) is False


def test_schedule_deactivates_when_creator_gone(tenant_a, user_in):
    """A schedule whose creator was deleted (created_by=NULL) must deactivate
    rather than silently fire empty, undelivered runs every cadence window."""
    with schema_context(tenant_a.schema_name):
        director = user_in(tenant_a, roles=[Role.DIRECTOR])
        now = timezone.localtime(timezone.now())
        sched = ReportSchedule.objects.create(
            report=Report.objects.get(key="enrollment"),
            created_by=director,
            cadence=ReportSchedule.Cadence.WEEKLY,
            weekday=now.weekday(),
            hour=now.hour,
            format="pdf",
        )
        ReportSchedule.objects.filter(pk=sched.pk).update(created_by=None)
        fired = services.run_due_schedules(now=now)
        sched.refresh_from_db()
        run_count = ReportRun.objects.filter(report__key="enrollment").count()
    assert fired == 0
    assert sched.is_active is False
    assert run_count == 0


def test_scheduled_run_carries_and_delivers_to_recipients(tenant_a, user_in, monkeypatch):
    """recipient_ids configured on a schedule are copied onto the run and the ready
    notification is delivered to them (plus the requester), not ignored."""
    delivered: list[int] = []

    def _capture(**kwargs):
        delivered.append(kwargs["recipient_id"])

    monkeypatch.setattr("apps.notifications.services.dispatch", _capture)
    _patch_s3(monkeypatch)

    with schema_context(tenant_a.schema_name):
        branch = BranchFactory(name="Scheduled Delivery", slug="scheduled-delivery")
        creator = user_in(tenant_a, roles=[Role.DIRECTOR], branch=branch)
        extra = user_in(tenant_a, roles=[Role.HEAD_OF_DEPT], branch=branch)
        now = timezone.localtime(timezone.now())
        sched = ReportSchedule.objects.create(
            report=Report.objects.get(key="enrollment"),
            created_by=creator,
            cadence=ReportSchedule.Cadence.WEEKLY,
            weekday=now.weekday(),
            hour=now.hour,
            format="pdf",
            recipient_ids=[extra.pk],
            params={"branch_id": branch.pk},
        )
        run = services.fire_schedule(sched, now=now)
        assert run.recipient_ids == [extra.pk]  # copied from the schedule

        # _notify_ready (called by the build task) delivers to requester + recipients.
        run.status = ReportRun.Status.DONE
        run.s3_key = f"{tenant_a.schema_name}/reports/{run.pk}.pdf"
        run.save(update_fields=["status", "s3_key"])
        services._notify_ready(run)

    assert set(delivered) == {creator.pk, extra.pk}


@pytest.mark.skipif(not _HAS_OPENPYXL, reason="openpyxl not installed")
def test_xlsx_export_neutralizes_formula_injection(tenant_a):
    """A user-controlled cell beginning with '=' must be written as text, not an
    active formula, in the XLSX export (CSV/Excel formula injection)."""
    import io

    from openpyxl import load_workbook

    gen = get_generator("enrollment")
    data = {"columns": ["student"], "rows": [{"student": '=HYPERLINK("http://evil","x")'}]}
    raw = gen.render_xlsx(data)
    ws = load_workbook(io.BytesIO(raw)).active
    assert ws.cell(row=2, column=1).value.startswith("'=")  # apostrophe-escaped


def test_safe_cell_neutralizes_formula_prefixes():
    """Unit cover for safe_cell regardless of whether openpyxl is installed."""
    from apps.reports.generators.base import safe_cell

    for payload in ("=1+1", "+1", "-1", "@SUM(A1)"):
        assert safe_cell(payload) == "'" + payload
    assert safe_cell("Ali Valiyev") == "Ali Valiyev"  # ordinary text untouched
    assert safe_cell(42) == 42  # non-strings pass through


# --------------------------------------------------------------------------- #
# API: cross-tenant isolation + permissions + query-count (D4-LB-5)
# --------------------------------------------------------------------------- #
def test_runs_cross_tenant_isolation(tenant_a, tenant_b, user_in, client_for):
    from apps.auth.services import issue_token

    with schema_context(tenant_a.schema_name):
        a_user = user_in(tenant_a, roles=[Role.DIRECTOR])
        ReportRun.objects.create(
            report=Report.objects.get(key="enrollment"),
            requested_by=a_user,
            format="pdf",
            status=ReportRun.Status.DONE,
        )
        access = issue_token(a_user)["access"]
    # A tenant-A token used against tenant-B's host must 401 (TD-1 tenant_mismatch).
    client = client_for(tenant_b)
    client.credentials(HTTP_AUTHORIZATION=f"Bearer {access}")
    resp = client.get("/api/v1/reports/runs/")
    assert resp.status_code == 401


def test_runs_list_query_count(tenant_a, user_in, as_user, django_assert_max_num_queries):
    with schema_context(tenant_a.schema_name):
        director = user_in(tenant_a, roles=[Role.DIRECTOR])
        for _ in range(5):
            ReportRun.objects.create(
                report=Report.objects.get(key="enrollment"),
                requested_by=director,
                format="pdf",
                status=ReportRun.Status.QUEUED,
            )
    client = as_user(tenant_a, director)
    with django_assert_max_num_queries(12):
        resp = client.get("/api/v1/reports/runs/")
    assert resp.status_code == 200
    # Standard envelope (StandardEnvelopeRenderer) — reports matches every other app:
    # {"success": true, "data": [...], "pagination": {"total", ...}}, NOT DRF's
    # bare {"count", "results"}.
    body = resp.json()
    assert body["success"] is True
    assert body["pagination"]["total"] == 5
    assert len(body["data"]) == 5
    assert body["pagination"]["page"] == 1


def test_run_detail_and_datetime_envelope(tenant_a, user_in, as_user):
    """A single run: enveloped {success,data} AND timestamps rendered UTC +00:00
    (matching the layered presenters), not DRF's localized +05:00 / Z."""
    with schema_context(tenant_a.schema_name):
        director = user_in(tenant_a, roles=[Role.DIRECTOR])
        run = ReportRun.objects.create(
            report=Report.objects.get(key="enrollment"),
            requested_by=director,
            format="pdf",
            status=ReportRun.Status.QUEUED,
        )
    resp = as_user(tenant_a, director).get(f"/api/v1/reports/runs/{run.pk}/")
    assert resp.status_code == 200
    body = resp.json()
    assert body["success"] is True
    assert body["data"]["id"] == run.pk
    assert body["data"]["created_at"].endswith("+00:00")


# --------------------------------------------------------------------------- #
# Render path — SKIP when the optional lib is absent (assert data layer instead)
# --------------------------------------------------------------------------- #
@pytest.mark.skipif(not _HAS_WEASYPRINT, reason="weasyprint native libs unavailable (CI/Linux runs it)")
def test_pdf_render_real(tenant_a, user_in):
    with schema_context(tenant_a.schema_name):
        director = user_in(tenant_a, roles=[Role.DIRECTOR])
        StudentProfileFactory(status=StudentProfile.Status.ACTIVE)
        gen = get_generator("enrollment")
        data = gen.collect({}, user=director, roles={Role.DIRECTOR})
        pdf = gen.render_pdf(data, locale="uz")
    assert pdf.startswith(b"%PDF")


@pytest.mark.skipif(not _HAS_OPENPYXL, reason="openpyxl unavailable (CI runs it)")
def test_xlsx_render_real(tenant_a, user_in):
    with schema_context(tenant_a.schema_name):
        director = user_in(tenant_a, roles=[Role.DIRECTOR])
        StudentProfileFactory(status=StudentProfile.Status.ACTIVE)
        gen = get_generator("enrollment")
        data = gen.collect({}, user=director, roles={Role.DIRECTOR})
        xlsx = gen.render_xlsx(data)
    # xlsx is a zip container — magic bytes "PK".
    assert xlsx[:2] == b"PK"
