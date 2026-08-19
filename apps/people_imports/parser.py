"""Bounded CSV/TSV/XLSX parsing for reviewed people imports."""

from __future__ import annotations

import csv
import io
import zipfile
from collections.abc import Iterable
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from django.utils.translation import gettext_lazy as _

from core.exceptions import ValidationException

MAX_IMPORT_BYTES = 4 * 1024 * 1024
MAX_IMPORT_ROWS = 2_000
MAX_WORKBOOK_EXPANDED_BYTES = 32 * 1024 * 1024
MAX_WORKBOOK_FILES = 2_000
ALLOWED_EXTENSIONS = frozenset({".csv", ".tsv", ".xlsx"})


@dataclass(frozen=True)
class ParsedPeopleFile:
    rows: list[dict[str, str]]
    headers: tuple[str, ...]
    sheet_name: str = ""


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat() if value.time() == datetime.min.time() else value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _read_bounded(file_obj) -> bytes:
    declared = getattr(file_obj, "size", None)
    if declared is not None and declared > MAX_IMPORT_BYTES:
        raise ValidationException(
            _("Import files may not exceed 4 MB."),
            code="file_too_large",
            fields={"file": [_("Choose a file no larger than 4 MB.")]},
        )
    content = file_obj.read(MAX_IMPORT_BYTES + 1)
    if len(content) > MAX_IMPORT_BYTES:
        raise ValidationException(
            _("Import files may not exceed 4 MB."),
            code="file_too_large",
            fields={"file": [_("Choose a file no larger than 4 MB.")]},
        )
    if not content:
        raise ValidationException(
            _("The selected file is empty."),
            code="empty_file",
            fields={"file": [_("Choose a file containing a header row and at least one person.")]},
        )
    return content


def _matrix_to_rows(matrix: Iterable[Iterable[Any]]) -> tuple[list[dict[str, str]], tuple[str, ...]]:
    iterator = iter(matrix)
    headers: tuple[str, ...] = ()
    for candidate in iterator:
        header_values = tuple(_cell_text(value) for value in candidate)
        if any(header_values):
            headers = header_values
            break
    if not headers:
        raise ValidationException(
            _("No header row was found."),
            code="missing_headers",
            fields={"file": [_("The first non-empty row must contain column names.")]},
        )

    display_headers: list[str] = []
    seen: dict[str, int] = {}
    for index, value in enumerate(headers, start=1):
        base = value or f"Column {index}"
        seen[base] = seen.get(base, 0) + 1
        display_headers.append(base if seen[base] == 1 else f"{base} ({seen[base]})")

    rows: list[dict[str, str]] = []
    for source_row, candidate in enumerate(iterator, start=2):
        row_values = [_cell_text(value) for value in candidate]
        if not any(row_values):
            continue
        if len(rows) >= MAX_IMPORT_ROWS:
            raise ValidationException(
                _("This file contains more than 2,000 people."),
                code="too_many_rows",
                fields={"file": [_("Split the file into imports of 2,000 rows or fewer.")]},
            )
        row_values.extend([""] * max(0, len(display_headers) - len(row_values)))
        row = {header: row_values[index] for index, header in enumerate(display_headers)}
        row["__source_row__"] = str(source_row)
        rows.append(row)

    if not rows:
        raise ValidationException(
            _("No people were found below the header row."),
            code="empty_import",
            fields={"file": [_("Add at least one populated row below the headers.")]},
        )
    return rows, tuple(display_headers)


def _parse_delimited(content: bytes, *, extension: str) -> ParsedPeopleFile:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        raise ValidationException(
            _("Text imports must use UTF-8 encoding."),
            code="invalid_encoding",
            fields={"file": [_("Export the spreadsheet as UTF-8 CSV and try again.")]},
        ) from None
    delimiter = "\t" if extension == ".tsv" else ","
    if extension == ".csv":
        try:
            delimiter = csv.Sniffer().sniff(text[:8_192], delimiters=",;\t|").delimiter
        except csv.Error:
            delimiter = ","
    try:
        rows, headers = _matrix_to_rows(csv.reader(io.StringIO(text), delimiter=delimiter))
    except csv.Error as exc:
        raise ValidationException(
            _("The delimited file could not be read."),
            code="invalid_file",
            fields={"file": [_("Export a standard CSV or TSV file and try again.")]},
        ) from exc
    return ParsedPeopleFile(rows=rows, headers=headers)


def _validate_workbook_archive(content: bytes) -> None:
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as archive:
            members = archive.infolist()
            expanded = sum(member.file_size for member in members)
            if len(members) > MAX_WORKBOOK_FILES or expanded > MAX_WORKBOOK_EXPANDED_BYTES:
                raise ValidationException(
                    _("The workbook expands beyond the safe import limit."),
                    code="workbook_too_large",
                    fields={"file": [_("Remove embedded content or split the workbook.")]},
                )
    except zipfile.BadZipFile:
        raise ValidationException(
            _("The selected workbook is not a valid XLSX file."),
            code="invalid_file",
            fields={"file": [_("Open and re-save the workbook as .xlsx, then try again.")]},
        ) from None


def _parse_xlsx(content: bytes) -> ParsedPeopleFile:
    _validate_workbook_archive(content)
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(
            io.BytesIO(content),
            read_only=True,
            data_only=True,
            keep_links=False,
        )
    except Exception as exc:
        raise ValidationException(
            _("The selected workbook could not be read."),
            code="invalid_file",
            fields={"file": [_("Open and re-save the workbook as .xlsx, then try again.")]},
        ) from exc
    try:
        for worksheet in workbook.worksheets:
            # A workbook can advertise all 16,384 Excel columns even when the
            # visible sheet is tiny. Keep the review surface intentionally
            # tabular and bounded before materialising any cells in memory.
            matrix = list(
                worksheet.iter_rows(
                    values_only=True,
                    max_row=MAX_IMPORT_ROWS + 2,
                    max_col=128,
                )
            )
            if not any(any(_cell_text(value) for value in row) for row in matrix):
                continue
            rows, headers = _matrix_to_rows(matrix)
            return ParsedPeopleFile(rows=rows, headers=headers, sheet_name=worksheet.title[:255])
    finally:
        workbook.close()
    raise ValidationException(
        _("The workbook does not contain a populated worksheet."),
        code="empty_import",
        fields={"file": [_("Add a header row and at least one person, then try again.")]},
    )


def parse_people_file(file_obj) -> ParsedPeopleFile:
    filename = Path(str(getattr(file_obj, "name", "") or "import")).name
    extension = Path(filename).suffix.lower()
    if extension not in ALLOWED_EXTENSIONS:
        raise ValidationException(
            _("Choose a CSV, TSV, or XLSX file."),
            code="unsupported_file_type",
            fields={"file": [_("Supported formats are .csv, .tsv, and .xlsx.")]},
        )
    content = _read_bounded(file_obj)
    if extension == ".xlsx":
        return _parse_xlsx(content)
    return _parse_delimited(content, extension=extension)
