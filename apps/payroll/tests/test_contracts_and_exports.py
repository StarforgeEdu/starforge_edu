from __future__ import annotations

import io
from types import SimpleNamespace

from openpyxl import load_workbook

from apps.payroll.openapi_contracts import PAYROLL_SCHEMAS
from apps.payroll.presenters import export_to_dict
from apps.payroll.services import _EXPORT_COLUMNS, _render_pdf, _render_xlsx
from apps.payroll.urls import urlpatterns
from core.openapi import _validate_view_contract
from core.openapi_contracts import get_openapi_contract


def _row(**overrides):
    row = {
        "payslip": "PAY-1",
        "teacher_code": "T-1",
        "teacher_name": "Teacher",
        "branch": "Central",
        "department": "Languages",
        "method": "flat_monthly",
        "base_amount_uzs": "100.00",
        "bonus_amount_uzs": "0.00",
        "deduction_amount_uzs": "0.00",
        "net_amount_uzs": "100.00",
        "paid_amount_uzs": "0.00",
        "outstanding_amount_uzs": "100.00",
        "currency": "UZS",
        "payment_state": "unpaid",
    }
    row.update(overrides)
    return row


def test_xlsx_export_neutralizes_spreadsheet_formula_cells():
    payload = _render_xlsx([_row(teacher_name='=HYPERLINK("https://attacker.invalid","open")')])
    workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=False)
    sheet = workbook["Payroll"]
    values = list(sheet.iter_rows(values_only=False))
    teacher_column = _EXPORT_COLUMNS.index("teacher_name")
    cell = values[1][teacher_column]
    assert cell.data_type != "f"
    assert str(cell.value).startswith("'")


def test_pdf_export_html_escapes_untrusted_snapshot_text(monkeypatch):
    captured: dict[str, str] = {}

    class FakeHTML:
        def __init__(self, *, string: str):
            captured["html"] = string

        def write_pdf(self):
            return b"safe-pdf"

    monkeypatch.setattr("weasyprint.HTML", FakeHTML)
    export = SimpleNamespace(
        period=SimpleNamespace(
            label="<img src=x onerror=alert(1)>",
            period_start=SimpleNamespace(isoformat=lambda: "2026-06-01"),
            period_end=SimpleNamespace(isoformat=lambda: "2026-06-30"),
        )
    )
    result = _render_pdf(export, [_row(teacher_name="<script>alert(1)</script>")])
    assert result == b"safe-pdf"
    assert "<script>" not in captured["html"]
    assert "&lt;script&gt;" in captured["html"]
    assert "onerror=alert" in captured["html"]
    assert "<img" not in captured["html"]


def test_export_collections_do_not_mint_signed_urls(monkeypatch):
    calls: list[int] = []

    def fake_presign(export):
        calls.append(export.pk)
        return "https://files.invalid/signed"

    monkeypatch.setattr("apps.payroll.services.presign_export", fake_presign)
    export = SimpleNamespace(
        pk=9,
        period_id=3,
        format="xlsx",
        filters={},
        status="done",
        file_bytes=42,
        error_code="",
        created_at=None,
        started_at=None,
        finished_at=None,
    )
    collection_row = export_to_dict(export)
    assert collection_row["download_url"] is None
    assert calls == []
    detail = export_to_dict(export, include_download=True)
    assert detail["download_url"] == "https://files.invalid/signed"
    assert calls == [9]


def test_every_payroll_route_has_an_exact_explicit_contract():
    contracts = []
    for pattern in urlpatterns:
        contract = get_openapi_contract(pattern.callback)
        assert contract is not None, pattern.name
        assert contract.critical is True
        assert contract.exact_methods is True
        _validate_view_contract(
            contract=contract,
            callback=pattern.callback,
            path=contract.expected_path,
            name=pattern.name or "",
        )
        contracts.append(contract)
    assert len(contracts) == 24
    operation_ids = [operation.operation_id for contract in contracts for operation in contract.operations]
    assert None not in operation_ids
    assert len(operation_ids) == len(set(operation_ids))


def test_payroll_component_schemas_are_closed_named_contracts():
    forbidden_empty = [
        name
        for name, schema in PAYROLL_SCHEMAS.items()
        if schema == {"type": "object", "additionalProperties": False, "properties": {}}
    ]
    assert forbidden_empty == []
    assert "PaginatedResponse" not in repr(PAYROLL_SCHEMAS)
    assert PAYROLL_SCHEMAS["PayrollPeriodCreateRequest"]["additionalProperties"] is False
    assert PAYROLL_SCHEMAS["PayrollReconcileRequest"]["additionalProperties"] is False
    assert PAYROLL_SCHEMAS["PayrollAdjustmentCreateRequest"]["additionalProperties"] is False
    assert PAYROLL_SCHEMAS["PayrollPeriodListResponse"]["required"] == [
        "success",
        "data",
        "pagination",
    ]

    refs: set[str] = set()

    def walk(value):
        if isinstance(value, dict):
            ref = value.get("$ref")
            if isinstance(ref, str) and ref.startswith("#/components/schemas/"):
                refs.add(ref.rsplit("/", 1)[-1])
            for child in value.values():
                walk(child)
        elif isinstance(value, list):
            for child in value:
                walk(child)

    walk(PAYROLL_SCHEMAS)
    assert refs <= set(PAYROLL_SCHEMAS)
